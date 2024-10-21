import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def read_excel_files(depart_file, reference_file):
    """Lit les fichiers Excel et retourne les DataFrames correspondants."""
    depart_df = pd.read_excel(depart_file)
    reference_df = pd.read_excel(reference_file)
    return depart_df, reference_df


def find_match(value, reference_df):
    """Trouve la correspondance dans le fichier de référence."""
    for col in reference_df.columns:
        if col.startswith('I-Place'):
            match = reference_df[reference_df[col] == value]
            if not match.empty:
                place_col = col.replace('I-Place', 'Place')
                nbr_partants_col = col.replace('I-Place', 'NbrPartants')
                return match[place_col].values[0], match[nbr_partants_col].values[0]
    return np.nan, np.nan


def process_columns(depart_df, reference_df):
    """Traite les colonnes et crée le DataFrame résultant."""
    result_df = pd.DataFrame()
    for i in range(1, 5):
        i_place_col = f'I-Place-{i}'
        t_place_col = f'T-Place-{i}'
        place_col = f'Place-{i}'
        nbr_partants_col = f'NbrPartants-{i}'

        result_df[t_place_col] = depart_df[i_place_col]
        result_df[place_col], result_df[nbr_partants_col] = zip(
            *depart_df[i_place_col].apply(lambda x: find_match(x, reference_df)))

    return result_df


def reorder_columns(df):
    """Réorganise les colonnes selon le format souhaité."""
    column_order = []
    for i in range(1, 5):
        column_order.extend([f'T-Place-{i}', f'Place-{i}', f'NbrPartants-{i}'])

    df = df.reindex(columns=column_order)
    return df


def save_result(df, output_file):
    """Sauvegarde le DataFrame résultant dans un fichier Excel avec mise en forme améliorée."""
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    wb.save(output_file)
    print(f"Le fichier {output_file} a été créé avec succès et formaté.")


def main():
    depart_file = 'FICHE2.xls'
    reference_file = 'REF-LISTE.xls'
    output_file = 'resultat.xlsx'

    depart_df, reference_df = read_excel_files(depart_file, reference_file)
    result_df = process_columns(depart_df, reference_df)
    result_df = reorder_columns(result_df)
    save_result(result_df, output_file)


if __name__ == "__main__":
    main()
